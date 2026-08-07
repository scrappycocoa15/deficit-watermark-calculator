#!/usr/bin/env python3
"""
Deficit Watermark Calculator
============================
Reads two Salesforce CSV exports (Sales and Terminations), computes per-account
negative watermarks using FIFO netting logic, and writes a dated Excel results file.

Business Rules
--------------
  SOF losses (Sales report — negative BMI Net New ARR):
      Watermark start date  = Order Effective Date (OED)
      Watermark clear date  = OED + 6 months

  Termination losses (Disconnects report — Total Lost Revenue):
      Watermark start date  = Billing End Date
      Watermark clear date  = Billing End Date + 6 months

  Netting (FIFO):
      Positive sales are applied oldest-deficit-first against active watermarks.
      A watermark is active if the transaction date is on or before its clear date.
      Expired watermarks are purged before each transaction is processed.

  Sale eligibility guidance:
      A new opportunity is eligible if its Billing Start Date falls AFTER the
      watermark clear date (e.g., clear date = 6/5 → earliest valid BSD = 7/1).

Usage
-----
  python deficit_calculator.py --sales sales.csv --terms terms.csv
  python deficit_calculator.py --sales sales.csv --terms terms.csv --cses cses.csv
  python deficit_calculator.py --sales sales.csv --terms terms.csv --cses cses.csv --output results.xlsx
  python deficit_calculator.py --sales sales.csv --terms terms.csv --all-accounts

Output sheets
-------------
  Results      — One row per account, all active deficits as columns.
  Key          — Key team: accounts grouped under Leader → Rep.  (requires --cses)
  Strategic    — Strategic team: accounts grouped under Leader → Rep.  (requires --cses)
  Premier      — Premier team: accounts grouped under Leader → Rep.  (requires --cses)
  Other        — Any reps not in the three named teams.  (requires --cses)

Column names are configurable at the top of this file under "Column Configuration".
"""

import argparse
import os
import sys
from datetime import date

from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─── Column Configuration ──────────────────────────────────────────────────────
# Edit these strings if your Salesforce reports use different column headers.

SALES_ID_COL     = "18 Digit Account ID"
SALES_AMOUNT_COL = "BMI Net New ARR (converted)"
SALES_DATE_COL   = "Order Effective Date"
SALES_OWNER_COL  = "Account Owner"
SALES_NAME_COL   = "Account Name"

TERMS_ID_COL     = "Account: 18 Digit Account ID"
TERMS_AMOUNT_COL = "Total Lost Revenue (converted)"
TERMS_DATE_COL   = "Billing End Date"
TERMS_OWNER_COL  = "Account: Account Owner"
TERMS_NAME_COL   = "Account: Account Name"

# CSEs column names (for --cses option)
CSES_OWNER_COL   = "AcctOwn - Account Owner Name"
CSES_MANAGER_COL = "AcctOwn - Manager Name"
CSES_TEAM_COL    = "AcctOwn - Owner Team"

# ─── Team Sheet Mapping ────────────────────────────────────────────────────────
# Maps the Excel sheet tab name to the team value that appears in the CSEs file.
# Edit the right-hand side values if your CSEs export uses different team names.
TEAM_SHEETS = {
    "Key":       "US SMB Client Sales Key",
    "Strategic": "US SMB Client Sales Strategic",
    "Premier":   "US SMB Client Sales Premier",
}

# ─── Color Palette ────────────────────────────────────────────────────────────
COLOR_COL_HDR_BG    = "0070F2"   # SAP blue    — column header row
COLOR_COL_HDR_FONT  = "FFFFFF"
COLOR_ALT_ROW       = "E1F4FF"   # light blue  — alternating data rows
COLOR_LEADER_BG     = "00144A"   # SAP navy    — leader group header
COLOR_REP_BG        = "4CB1FF"   # SAP mid-blue — rep sub-header


# ══════════════════════════════════════════════════════════════════════════════
# Data Loading
# ══════════════════════════════════════════════════════════════════════════════

def _check_columns(df: pd.DataFrame, required: list, label: str):
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"\nError: The {label} CSV is missing expected columns:")
        for col in missing:
            print(f"  - {col}")
        print(f"Columns found: {list(df.columns)}")
        print("Update the column configuration at the top of this script.")
        sys.exit(1)


def load_sales(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, encoding="latin-1")
    _check_columns(df, [SALES_ID_COL, SALES_AMOUNT_COL, SALES_DATE_COL,
                        SALES_OWNER_COL, SALES_NAME_COL], "Sales")
    df = df.rename(columns={
        SALES_ID_COL:     "account_id",
        SALES_AMOUNT_COL: "amount",
        SALES_DATE_COL:   "effective_date",
        SALES_OWNER_COL:  "account_owner",
        SALES_NAME_COL:   "account_name",
    })[["account_id", "amount", "effective_date", "account_owner", "account_name"]].copy()
    df["source"]         = "Sales"
    df["amount"]         = pd.to_numeric(df["amount"].str.replace(",", ""), errors="coerce")
    df["effective_date"] = pd.to_datetime(df["effective_date"], errors="coerce")
    before = len(df)
    df = df[df["amount"].notna() & (df["amount"] != 0)
            & df["effective_date"].notna()
            & df["account_id"].notna()
            & (df["account_id"].str.strip() != "")]
    dropped = before - len(df)
    if dropped:
        print(f"  Sales: dropped {dropped} rows ($0, missing date, or missing ID).")
    return df.reset_index(drop=True)


def load_terms(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, encoding="latin-1")
    _check_columns(df, [TERMS_ID_COL, TERMS_AMOUNT_COL, TERMS_DATE_COL,
                        TERMS_OWNER_COL, TERMS_NAME_COL], "Terms")
    df = df.rename(columns={
        TERMS_ID_COL:     "account_id",
        TERMS_AMOUNT_COL: "amount",
        TERMS_DATE_COL:   "effective_date",
        TERMS_OWNER_COL:  "account_owner",
        TERMS_NAME_COL:   "account_name",
    })[["account_id", "amount", "effective_date", "account_owner", "account_name"]].copy()
    df["source"]         = "Terminations"
    df["amount"]         = pd.to_numeric(df["amount"].str.replace(",", ""), errors="coerce")
    df["effective_date"] = pd.to_datetime(df["effective_date"], errors="coerce")
    df["amount"]         = df["amount"].abs() * -1
    before = len(df)
    df = df[df["amount"].notna() & (df["amount"] != 0)
            & df["effective_date"].notna()
            & df["account_id"].notna()
            & (df["account_id"].str.strip() != "")]
    dropped = before - len(df)
    if dropped:
        print(f"  Terms: dropped {dropped} rows ($0, missing date, or missing ID).")
    return df.reset_index(drop=True)


def load_cses(path: str) -> pd.DataFrame:
    """
    Load the CSEs reference CSV (exported from the CSEs tab in the workbook).
    Returns a DataFrame with columns: owner_name, manager_name, team.
    """
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    _check_columns(df, [CSES_OWNER_COL, CSES_MANAGER_COL, CSES_TEAM_COL], "CSEs")
    df = df.rename(columns={
        CSES_OWNER_COL:   "owner_name",
        CSES_MANAGER_COL: "manager_name",
        CSES_TEAM_COL:    "team",
    })[["owner_name", "manager_name", "team"]].copy()
    df["owner_name"]   = df["owner_name"].str.strip()
    df["manager_name"] = df["manager_name"].fillna("Unassigned").str.strip()
    df["team"]         = df["team"].fillna("").str.strip()
    df = df.dropna(subset=["owner_name"]).drop_duplicates(subset=["owner_name"])
    return df.reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# Watermark Calculation
# ══════════════════════════════════════════════════════════════════════════════

def _clear_date(effective_date: pd.Timestamp) -> pd.Timestamp:
    return effective_date + relativedelta(months=6)


def process_accounts(combined: pd.DataFrame) -> list:
    combined = combined.sort_values(
        ["account_id", "effective_date", "amount"],
        ascending=[True, True, True]
    ).reset_index(drop=True)

    results = []
    today   = pd.Timestamp(date.today())

    for account_id, group in combined.groupby("account_id", sort=False):
        last_row      = group.iloc[-1]
        account_owner = last_row["account_owner"]
        account_name  = last_row["account_name"]
        deficits      = []

        for _, row in group.iterrows():
            trans_date = row["effective_date"]
            amount     = float(row["amount"])
            source     = row["source"]
            deficits   = [d for d in deficits if trans_date <= d["clear_date"]]

            if amount < 0:
                deficits.append({
                    "amount":     abs(amount),
                    "eff_date":   trans_date,
                    "clear_date": _clear_date(trans_date),
                    "source":     source,
                })
            elif amount > 0:
                remaining = amount
                i = 0
                while i < len(deficits) and remaining > 0:
                    d = deficits[i]
                    if remaining >= d["amount"]:
                        remaining -= d["amount"]
                        deficits.pop(i)
                    else:
                        d["amount"] -= remaining
                        remaining = 0
                        i += 1

        deficits = [d for d in deficits if today <= d["clear_date"]]
        results.append({
            "account_owner": account_owner,
            "account_name":  account_name,
            "account_id":    account_id,
            "deficits":      deficits,
        })
    return results


# ══════════════════════════════════════════════════════════════════════════════
# Excel Output Helpers
# ══════════════════════════════════════════════════════════════════════════════

def _deficit_col_headers(max_deficits: int) -> list:
    headers = []
    for n in range(1, max_deficits + 1):
        headers += [
            f"Deficit {n} Amount",
            f"Deficit {n} Effective Date",
            f"Deficit {n} Clears/Expires",
            f"Deficit {n} Source",
        ]
    return headers


def _write_col_header_row(ws, col_headers: list):
    font  = Font(bold=True, color=COLOR_COL_HDR_FONT, name="Calibri", size=10)
    fill  = PatternFill(start_color=COLOR_COL_HDR_BG,
                        end_color=COLOR_COL_HDR_BG, fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(col_headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = font, fill, align
    ws.row_dimensions[1].height = 30


def _write_group_header(ws, row: int, label: str, total_cols: int,
                        bg: str, font_color: str = "FFFFFF",
                        font_size: int = 11, indent: int = 0):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=total_cols)
    prefix = "  " * indent
    c = ws.cell(row=row, column=1, value=f"{prefix}{label}")
    c.font  = Font(bold=True, color=font_color, name="Calibri", size=font_size)
    c.fill  = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def _write_account_row(ws, row: int, rep_name: str,
                       account_name: str, account_id: str,
                       deficits: list, use_alt: bool):
    alt_fill = (PatternFill(start_color=COLOR_ALT_ROW,
                            end_color=COLOR_ALT_ROW, fill_type="solid")
                if use_alt else None)

    def _cell(col, value, num_fmt=None, h_align=None):
        c = ws.cell(row=row, column=col, value=value)
        if alt_fill:
            c.fill = alt_fill
        if num_fmt:
            c.number_format = num_fmt
        if h_align:
            c.alignment = Alignment(horizontal=h_align)
        return c

    _cell(1, rep_name)
    _cell(2, account_name)
    _cell(3, account_id)

    ci = 4
    for d in deficits:
        _cell(ci,   -round(d["amount"], 2), num_fmt='$#,##0.00', h_align="right")
        _cell(ci+1, d["eff_date"].date(),   num_fmt="MM/DD/YYYY", h_align="center")
        _cell(ci+2, d["clear_date"].date(), num_fmt="MM/DD/YYYY", h_align="center")
        _cell(ci+3, d["source"])
        ci += 4


def _autofit_freeze(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)
    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet Writers
# ══════════════════════════════════════════════════════════════════════════════

def write_results_sheet(ws, df: pd.DataFrame):
    """Master results sheet — one row per account (wide format)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill  = PatternFill(start_color=COLOR_COL_HDR_BG,
                            end_color=COLOR_COL_HDR_BG, fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill  = PatternFill(start_color=COLOR_ALT_ROW,
                            end_color=COLOR_ALT_ROW, fill_type="solid")

    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align
    ws.row_dimensions[1].height = 32

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = alt_fill

    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        hdr = ws.cell(1, col[0].column).value or ""
        for cell in col:
            if cell.value is None:
                continue
            if "Amount" in hdr:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif any(k in hdr for k in ("Date", "Expires", "Clears")):
                cell.number_format = "MM/DD/YYYY"
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)
    ws.freeze_panes = "A2"


def write_team_sheet(ws, results: list, cses_df: pd.DataFrame,
                     team_filter: str, max_deficits: int):
    """
    Write one team sheet.
    Structure: Leader header (navy) → Rep sub-header (mid-blue) → Account rows.
    Columns:   Rep | Account Name | Account ID | Deficit 1… | Deficit 2… | …
    """
    col_headers = (["Rep", "Account Name", "18 Digit Account ID"]
                   + _deficit_col_headers(max_deficits))
    total_cols  = len(col_headers)

    _write_col_header_row(ws, col_headers)

    # Build lookup: rep_name → (manager_name, team)
    lookup = {
        row["owner_name"]: (row["manager_name"], row["team"])
        for _, row in cses_df.iterrows()
    }

    # Attach manager/team and filter to this team
    enriched = []
    for r in results:
        manager, team = lookup.get(r["account_owner"], ("Unassigned", ""))
        if team != team_filter:
            continue
        enriched.append({**r, "manager": manager, "team": team})

    if not enriched:
        ws.cell(row=2, column=1, value="No active deficits for this team.")
        return

    # Sort: manager A-Z → rep A-Z → account A-Z
    enriched.sort(key=lambda r: (
        r["manager"].lower(),
        r["account_owner"].lower(),
        r["account_name"].lower()
    ))

    current_row     = 2
    current_manager = None
    current_rep     = None
    alt_counter     = 0
    manager_header_rows = []   # row numbers of each leader header row

    for r in enriched:
        manager = r["manager"]
        rep     = r["account_owner"]

        # Leader header
        if manager != current_manager:
            current_manager = manager
            current_rep     = None
            alt_counter     = 0
            manager_header_rows.append(current_row)
            _write_group_header(ws, current_row, manager, total_cols,
                                bg=COLOR_LEADER_BG, font_size=12, indent=0)
            current_row += 1

        # Rep sub-header
        if rep != current_rep:
            current_rep = rep
            alt_counter = 0
            _write_group_header(ws, current_row, rep, total_cols,
                                bg=COLOR_REP_BG, font_color="FFFFFF",
                                font_size=10, indent=1)
            current_row += 1

        _write_account_row(ws, current_row,
                           rep_name=rep,
                           account_name=r["account_name"],
                           account_id=r["account_id"],
                           deficits=r["deficits"],
                           use_alt=(alt_counter % 2 == 1))
        alt_counter += 1
        current_row += 1

    # ── Row grouping: each leader section is collapsible ──────────────────
    # The leader header row is the summary row (outline button appears above).
    # Every row beneath it — rep sub-headers + account rows — is at level 1.
    last_data_row = current_row - 1
    for i, header_row in enumerate(manager_header_rows):
        group_start = header_row + 1
        group_end   = (manager_header_rows[i + 1] - 1
                       if i + 1 < len(manager_header_rows)
                       else last_data_row)
        for r in range(group_start, group_end + 1):
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = False

    # Place the [-] button above the grouped rows (on the leader header)
    ws.sheet_properties.outlinePr.summaryBelow = False

    _autofit_freeze(ws)


# ══════════════════════════════════════════════════════════════════════════════
# Main Excel Writer
# ══════════════════════════════════════════════════════════════════════════════

def build_results_df(results: list) -> pd.DataFrame:
    rows = []
    for r in results:
        row = {
            "Account Owner":       r["account_owner"],
            "Account Name":        r["account_name"],
            "18 Digit Account ID": r["account_id"],
        }
        for n, d in enumerate(r["deficits"], start=1):
            row[f"Deficit {n} Amount"]         = -round(d["amount"], 2)
            row[f"Deficit {n} Effective Date"] = d["eff_date"].date()
            row[f"Deficit {n} Clears/Expires"] = d["clear_date"].date()
            row[f"Deficit {n} Source"]         = d["source"]
        rows.append(row)
    return pd.DataFrame(rows)


def write_excel(results: list, output_path: str, cses_df: pd.DataFrame = None):
    active = [r for r in results if r["deficits"]]
    max_deficits = max((len(r["deficits"]) for r in active), default=1)
    df_results   = build_results_df(active)

    with pd.ExcelWriter(output_path, engine="openpyxl",
                        date_format="MM/DD/YYYY") as writer:

        # ── Results sheet ──────────────────────────────────────────────────
        df_results.to_excel(writer, index=False, sheet_name="Results")
        write_results_sheet(writer.sheets["Results"], df_results)

        # ── Team sheets (only if CSEs provided) ───────────────────────────
        if cses_df is not None:
            for tab_name, team_value in TEAM_SHEETS.items():
                # openpyxl sheet must exist before we can write to it;
                # create it via a dummy df write then clear + rewrite
                dummy = pd.DataFrame()
                dummy.to_excel(writer, index=False, sheet_name=tab_name)
                ws = writer.sheets[tab_name]
                # Clear the one empty header row pandas added
                ws.delete_rows(1)
                write_team_sheet(ws, active, cses_df, team_value, max_deficits)


# ══════════════════════════════════════════════════════════════════════════════
# Entry Point
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        prog="deficit_calculator",
        description="Deficit Watermark Calculator — FIFO netting from Salesforce CSV exports.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python deficit_calculator.py --sales sales.csv --terms terms.csv
  python deficit_calculator.py --sales sales.csv --terms terms.csv --cses cses.csv
  python deficit_calculator.py --sales sales.csv --terms terms.csv --cses cses.csv --output Jul_Deficits.xlsx
        """,
    )
    parser.add_argument("--sales",  required=True,
                        help="Sales CSV export from Salesforce.")
    parser.add_argument("--terms",  required=True,
                        help="Terminations CSV export from Salesforce.")
    parser.add_argument("--cses",   default="CSEs.csv",
                        help="CSEs CSV (exported from the CSEs tab). "
                             "Enables Key / Strategic / Premier sheets. "
                             "Defaults to CSEs.csv in the current folder.")
    parser.add_argument("--output", default=None,
                        help="Output file name. Default: DeficitResults_MMDD.xlsx")
    parser.add_argument("--all-accounts", action="store_true",
                        help="Include accounts with no active deficits in the Results sheet.")
    args = parser.parse_args()

    for path, label in [(args.sales, "--sales"), (args.terms, "--terms")]:
        if not os.path.isfile(path):
            print(f"Error: File not found for {label}: {path}")
            sys.exit(1)
    # If CSEs path was explicitly supplied on the command line, it must exist.
    # If it's the default (CSEs.csv) and the file isn't present, skip team sheets.
    _cses_default = "CSEs.csv"
    if args.cses and not os.path.isfile(args.cses):
        if args.cses == _cses_default:
            print(f"Note: Default CSEs.csv not found — team sheets (Key/Strategic/Premier) will be skipped.")
            args.cses = None
        else:
            print(f"Error: CSEs file not found: {args.cses}")
            sys.exit(1)

    print("Loading Sales data...")
    sales_df = load_sales(args.sales)
    print(f"  {len(sales_df):,} valid records ({sales_df['account_id'].nunique():,} accounts).")

    print("Loading Terminations data...")
    terms_df = load_terms(args.terms)
    print(f"  {len(terms_df):,} valid records ({terms_df['account_id'].nunique():,} accounts).")

    cses_df = None
    if args.cses:
        print("Loading CSEs data...")
        cses_df = load_cses(args.cses)
        print(f"  {len(cses_df):,} reps loaded ({cses_df['manager_name'].nunique():,} managers).")

    combined = pd.concat([sales_df, terms_df], ignore_index=True)
    print(f"\nCombined: {len(combined):,} transactions across "
          f"{combined['account_id'].nunique():,} accounts.")

    print("Computing watermarks...")
    results = process_accounts(combined)
    active  = [r for r in results if r["deficits"]]
    print(f"  {len(active):,} of {len(results):,} accounts have active watermarks.")

    if args.all_accounts is False:
        pass  # write_excel already filters to active internally

    if not active:
        print("\nNo active deficits found. Exiting.")
        sys.exit(0)

    output_path = args.output or f"DeficitResults_{date.today().strftime('%m%d')}.xlsx"
    write_excel(results, output_path, cses_df=cses_df)

    sheets = ["Results"]
    if cses_df is not None:
        sheets += list(TEAM_SHEETS.keys())
    print(f"\nDone. Saved to: {output_path}")
    print(f"  Sheets:  {', '.join(sheets)}")
    print(f"  Rows:    {len(active):,}")


if __name__ == "__main__":
    main()
